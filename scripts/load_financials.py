#!/usr/bin/env python3
"""
Financial Data Loader - Parse SystematIQ Excel workbooks into Supabase.

Usage:
    python3 load_financials.py --file "/path/to/Pies n Thighs 2025 Financial Reports- Final.xlsx"
    python3 load_financials.py --file "/path/to/workbook.xlsx" --dry-run
    python3 load_financials.py --file "/path/to/workbook.xlsx" --location kent_ave

Reads the P&L, Balance Sheet, and Cash Flow sheets from SystematIQ workbooks.
Also reads the Periods sheet to populate fiscal_periods.
Uses the 4-4-5 fiscal calendar.
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
load_dotenv(SCRIPT_DIR.parent / ".env.toast")

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://zxqtclvljxvdxsnmsqka.supabase.co"
)
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

BATCH_SIZE = 200

# ---------------------------------------------------------------------------
# Sheet configuration
# ---------------------------------------------------------------------------

# Map sheet names to statement types
SHEET_MAP = {
    "PnT P&L Model Willamsburg": {"statement": "pnl", "location": "kent_ave"},
    "PnT P&L Model Williamsburg": {"statement": "pnl", "location": "kent_ave"},
    "Balance Sheet": {"statement": "balance_sheet", "location": "kent_ave"},
    "Cash Flow Statement": {"statement": "cash_flow", "location": "kent_ave"},
}

# P&L section mapping: row labels -> (section, is_channel_breakdown)
# The P&L has a specific structure where some sections break out by channel
PNL_SECTIONS = {
    # Revenue sections with channel breakdown
    "Transaction Count": ("transaction_count", True),
    "Transactions": ("transaction_count", True),
    "Avg Ticket": ("avg_ticket", True),
    "Average Ticket": ("avg_ticket", True),
    "Gross Sales": ("gross_sales", True),
    "Sales Adjustments": ("sales_adjustments", True),
    "Net Sales": ("net_sales", True),
    # Cost sections (no channel breakdown)
    "COGS": ("cogs", False),
    "Cost of Goods Sold": ("cogs", False),
    "Product Margin": ("product_margin", False),
    "Store Labor": ("labor", False),
    "Labor": ("labor", False),
    "Occupancy": ("occupancy", False),
    "Selling": ("selling", False),
    "Selling Expenses": ("selling", False),
    "Other Store": ("other_store", False),
    "Other Store Expenses": ("other_store", False),
    "4-Wall EBITDA": ("four_wall_ebitda", False),
    "G&A": ("gna", False),
    "General & Administrative": ("gna", False),
    "Operating Income": ("operating_income", False),
    "Other Expenses": ("other_expenses", False),
    "Other Income/Expenses": ("other_expenses", False),
    "Net Income": ("net_income", False),
}

# Channel name normalization
CHANNEL_MAP = {
    "Dine In": "dine_in",
    "Dine-In": "dine_in",
    "Takeout": "takeout",
    "Take Out": "takeout",
    "In-Store": "in_store",
    "In Store": "in_store",
    "DoorDash": "doordash",
    "GrubHub": "grubhub",
    "Grubhub": "grubhub",
    "UberEats": "ubereats",
    "Uber Eats": "ubereats",
    "3rd Party": "3pd_total",
    "Catering": "catering",
    "Goldbelly & Wholesale": "goldbelly_wholesale",
    "Goldbelly": "goldbelly_wholesale",
    "Total": "total",
}

# 4-4-5 quarter mapping: period -> quarter
def period_to_quarter(period_num: int) -> int:
    """Map period number (1-12) to quarter (1-4) in 4-4-5 calendar."""
    if period_num <= 3:
        return 1
    elif period_num <= 6:
        return 2
    elif period_num <= 9:
        return 3
    else:
        return 4

def period_weeks(period_num: int) -> int:
    """Return number of weeks for a period in 4-4-5 calendar."""
    # Pattern: 4-4-5 repeating (periods 3, 6, 9, 12 are 5-week periods)
    return 5 if period_num % 3 == 0 else 4


# ---------------------------------------------------------------------------
# Period parsing
# ---------------------------------------------------------------------------

def parse_period_label(label: str) -> tuple[int, int] | None:
    """
    Parse a period label like "P1'25" or "P12'24" into (fiscal_year, period_number).
    Returns (2025, 1) for "P1'25" etc.
    """
    # Match patterns like P1'25, P12'24, P1 '25
    m = re.match(r"P(\d{1,2})\s*['']\s*(\d{2})", label.strip())
    if m:
        period = int(m.group(1))
        year = 2000 + int(m.group(2))
        return (year, period)
    return None


def find_period_columns(ws, header_row: int = 1) -> list[tuple[int, int, int]]:
    """
    Scan the header row to find period columns.
    Returns list of (col_index, fiscal_year, period_number).
    """
    periods = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        val = str(cell.value or "").strip()
        parsed = parse_period_label(val)
        if parsed:
            periods.append((col, parsed[0], parsed[1]))
    return periods


# ---------------------------------------------------------------------------
# Periods sheet parser
# ---------------------------------------------------------------------------

def load_periods_sheet(wb: openpyxl.Workbook) -> list[dict]:
    """
    Parse the Periods sheet to extract fiscal period definitions.
    Returns list of fiscal_periods rows.
    """
    ws = None
    for name in ["Periods", "Period", "Fiscal Periods", "Calendar"]:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        print("  Warning: No Periods sheet found, skipping fiscal_periods", file=sys.stderr)
        return []

    periods = []
    # Look for rows with period data (year, period#, start_date, end_date)
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = [cell.value for cell in row]
        # Try to identify period rows - look for period labels
        for i, val in enumerate(values):
            if val and isinstance(val, str):
                parsed = parse_period_label(val)
                if parsed:
                    year, pnum = parsed
                    # Look for dates in the same row
                    start_date = None
                    end_date = None
                    for j, v in enumerate(values):
                        if isinstance(v, datetime):
                            if start_date is None:
                                start_date = v.strftime("%Y-%m-%d")
                            else:
                                end_date = v.strftime("%Y-%m-%d")
                    if start_date and end_date:
                        periods.append({
                            "fiscal_year": year,
                            "period_number": pnum,
                            "period_label": f"P{pnum}'{str(year)[2:]}",
                            "quarter": period_to_quarter(pnum),
                            "num_weeks": period_weeks(pnum),
                            "start_date": start_date,
                            "end_date": end_date,
                        })
                    break
    return periods


# ---------------------------------------------------------------------------
# P&L sheet parser
# ---------------------------------------------------------------------------

def load_pnl_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    location_id: str,
) -> list[dict]:
    """
    Parse a P&L sheet and extract line items with per-period amounts.
    Returns list of financials rows.
    """
    if sheet_name not in wb.sheetnames:
        # Try fuzzy match
        for name in wb.sheetnames:
            if "P&L" in name or "PnL" in name:
                sheet_name = name
                break
        else:
            print(f"  Warning: Sheet '{sheet_name}' not found", file=sys.stderr)
            return []

    ws = wb[sheet_name]
    print(f"  Parsing sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)", file=sys.stderr)

    # Find period columns in the header
    # Try first few rows for the header
    period_cols = []
    for header_row in range(1, 6):
        period_cols = find_period_columns(ws, header_row)
        if period_cols:
            print(f"  Found {len(period_cols)} period columns in row {header_row}", file=sys.stderr)
            break

    if not period_cols:
        print(f"  Warning: No period columns found in {sheet_name}", file=sys.stderr)
        return []

    rows = []
    current_section = "unknown"
    is_channel_section = False
    data_start_row = header_row + 1 if period_cols else 2

    for row_num in range(data_start_row, ws.max_row + 1):
        # Column A (or first column) typically has the line item name
        label_cell = ws.cell(row=row_num, column=1).value
        if label_cell is None:
            continue
        label = str(label_cell).strip()
        if not label:
            continue

        # Check if this is a section header
        section_match = None
        for section_name, (section_key, has_channels) in PNL_SECTIONS.items():
            if label.lower() == section_name.lower() or label.lower().startswith(section_name.lower()):
                section_match = (section_key, has_channels)
                break

        if section_match:
            current_section = section_match[0]
            is_channel_section = section_match[1]
            # Section headers might also have values (like "Total Net Sales")
            # Check if there are numeric values in this row
            has_values = False
            for col, year, period in period_cols:
                val = ws.cell(row=row_num, column=col).value
                if val is not None and isinstance(val, (int, float)):
                    has_values = True
                    break
            if not has_values:
                continue

        # Extract channel from line item name if in a channel section
        channel = None
        line_item = label

        if is_channel_section:
            # Check if this line item matches a known channel
            for ch_name, ch_key in CHANNEL_MAP.items():
                if label == ch_name or label.startswith(ch_name):
                    channel = ch_key
                    break

        # Skip pure percentage rows (like "COGS %", "Labor %")
        if label.endswith("%") or label.endswith("% "):
            continue

        # Extract amounts for each period
        for col, fiscal_year, period_number in period_cols:
            val = ws.cell(row=row_num, column=col).value
            if val is None:
                continue
            try:
                amount = round(float(val), 2)
            except (ValueError, TypeError):
                continue

            # Skip zero amounts for cleaner data
            if amount == 0:
                continue

            rows.append({
                "fiscal_year": fiscal_year,
                "period_number": period_number,
                "location_id": location_id,
                "statement": "pnl",
                "section": current_section,
                "line_item": line_item,
                "channel": channel,
                "amount": amount,
            })

    return rows


# ---------------------------------------------------------------------------
# Balance Sheet / Cash Flow parser (simpler - no channel breakdown)
# ---------------------------------------------------------------------------

def load_simple_statement(
    wb: openpyxl.Workbook,
    sheet_name: str,
    statement_type: str,
    location_id: str,
) -> list[dict]:
    """
    Parse a Balance Sheet or Cash Flow sheet.
    These don't have channel breakdowns - just section/line_item/amount.
    """
    if sheet_name not in wb.sheetnames:
        # Try fuzzy match
        for name in wb.sheetnames:
            if statement_type == "balance_sheet" and "Balance" in name:
                sheet_name = name
                break
            elif statement_type == "cash_flow" and "Cash Flow" in name:
                sheet_name = name
                break
        else:
            print(f"  Warning: Sheet '{sheet_name}' not found", file=sys.stderr)
            return []

    ws = wb[sheet_name]
    print(f"  Parsing sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)", file=sys.stderr)

    # Find period columns
    period_cols = []
    for header_row in range(1, 6):
        period_cols = find_period_columns(ws, header_row)
        if period_cols:
            break

    if not period_cols:
        print(f"  Warning: No period columns found in {sheet_name}", file=sys.stderr)
        return []

    rows = []
    current_section = "other"
    data_start_row = header_row + 1

    # Section detection for balance sheet
    bs_sections = {
        "cash": "assets_cash",
        "receivable": "assets_receivable",
        "other current assets": "assets_current",
        "fixed assets": "assets_fixed",
        "total assets": "assets_total",
        "current liabilities": "liabilities_current",
        "long term liabilities": "liabilities_long_term",
        "total liabilities": "liabilities_total",
        "equity": "equity",
    }
    cf_sections = {
        "operating": "cash_ops",
        "investing": "cash_investing",
        "financing": "cash_financing",
    }
    section_map = bs_sections if statement_type == "balance_sheet" else cf_sections

    for row_num in range(data_start_row, ws.max_row + 1):
        label_cell = ws.cell(row=row_num, column=1).value
        if label_cell is None:
            continue
        label = str(label_cell).strip()
        if not label:
            continue

        # Check for section headers
        label_lower = label.lower()
        for key, section in section_map.items():
            if key in label_lower:
                current_section = section
                break

        # Skip percentage rows
        if label.endswith("%"):
            continue

        for col, fiscal_year, period_number in period_cols:
            val = ws.cell(row=row_num, column=col).value
            if val is None:
                continue
            try:
                amount = round(float(val), 2)
            except (ValueError, TypeError):
                continue

            if amount == 0:
                continue

            rows.append({
                "fiscal_year": fiscal_year,
                "period_number": period_number,
                "location_id": location_id,
                "statement": statement_type,
                "section": current_section,
                "line_item": label,
                "channel": None,
                "amount": amount,
            })

    return rows


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def supabase_headers() -> dict:
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal,resolution=merge-duplicates",
    }


def upsert_fiscal_periods(session: requests.Session, rows: list[dict]) -> int:
    """Upsert fiscal period rows."""
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/fiscal_periods"
    headers = supabase_headers()
    resp = session.post(url, json=rows, headers=headers, timeout=30)
    if resp.status_code in (200, 201):
        return len(rows)
    print(f"  Fiscal periods upsert error: {resp.status_code} - {resp.text[:300]}", file=sys.stderr)
    return 0


def upsert_financials(session: requests.Session, rows: list[dict]) -> int:
    """Upsert financial line item rows."""
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/financials"
    headers = supabase_headers()
    loaded = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        resp = session.post(url, json=batch, headers=headers, timeout=60)
        if resp.status_code in (200, 201):
            loaded += len(batch)
        else:
            print(
                f"  Financials upsert error (batch {i // BATCH_SIZE}): {resp.status_code} - {resp.text[:500]}",
                file=sys.stderr,
            )
    return loaded


def log_pipeline_run(
    session: requests.Session,
    status: str,
    rows_loaded: int,
    error_message: str | None = None,
    metadata: dict | None = None,
) -> None:
    url = f"{SUPABASE_URL}/rest/v1/pipeline_runs"
    headers = supabase_headers()
    headers["Prefer"] = "return=minimal"
    today = datetime.now().strftime("%Y-%m-%d")
    row = {
        "script_name": "load_financials",
        "run_date": today,
        "status": status,
        "rows_loaded": rows_loaded,
        "error_message": error_message,
        "metadata": json.dumps(metadata or {}),
    }
    session.post(url, json=[row], headers=headers, timeout=30)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Load SystematIQ financial workbook into Supabase")
    parser.add_argument("--file", required=True, help="Path to the Excel workbook")
    parser.add_argument("--location", default="kent_ave", help="Location ID (default: kent_ave)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    args = parser.parse_args()

    filepath = Path(args.file).expanduser().resolve()
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading workbook: {filepath.name}", file=sys.stderr)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    print(f"  Sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)

    # 1. Load fiscal periods
    periods = load_periods_sheet(wb)
    print(f"\n  Fiscal periods found: {len(periods)}", file=sys.stderr)

    # 2. Load P&L
    pnl_rows = []
    for sheet_name, config in SHEET_MAP.items():
        if config["statement"] == "pnl" and sheet_name in wb.sheetnames:
            pnl_rows = load_pnl_sheet(wb, sheet_name, args.location)
            break
    # If no exact match, try fuzzy
    if not pnl_rows:
        for name in wb.sheetnames:
            if "P&L" in name or "PnL" in name:
                pnl_rows = load_pnl_sheet(wb, name, args.location)
                break
    print(f"  P&L rows: {len(pnl_rows)}", file=sys.stderr)

    # 3. Load Balance Sheet
    bs_rows = load_simple_statement(wb, "Balance Sheet", "balance_sheet", args.location)
    print(f"  Balance Sheet rows: {len(bs_rows)}", file=sys.stderr)

    # 4. Load Cash Flow
    cf_rows = load_simple_statement(wb, "Cash Flow Statement", "cash_flow", args.location)
    print(f"  Cash Flow rows: {len(cf_rows)}", file=sys.stderr)

    all_financial_rows = pnl_rows + bs_rows + cf_rows
    print(f"\n  Total financial rows: {len(all_financial_rows)}", file=sys.stderr)

    if args.dry_run:
        # Show sample
        print("\n--- Sample P&L rows ---", file=sys.stderr)
        for row in pnl_rows[:10]:
            print(
                f"  P{row['period_number']}'{str(row['fiscal_year'])[2:]} | "
                f"{row['section']:20s} | {row['line_item']:30s} | "
                f"{'ch:' + (row['channel'] or '-'):15s} | "
                f"${row['amount']:>12,.2f}",
                file=sys.stderr,
            )
        print("\n--- Sample BS rows ---", file=sys.stderr)
        for row in bs_rows[:5]:
            print(
                f"  P{row['period_number']}'{str(row['fiscal_year'])[2:]} | "
                f"{row['section']:20s} | {row['line_item']:30s} | "
                f"${row['amount']:>12,.2f}",
                file=sys.stderr,
            )
        return

    # Write to Supabase
    session = requests.Session()

    if periods:
        loaded_periods = upsert_fiscal_periods(session, periods)
        print(f"  Upserted {loaded_periods} fiscal periods", file=sys.stderr)

    loaded_financials = upsert_financials(session, all_financial_rows)
    print(f"  Upserted {loaded_financials} financial rows", file=sys.stderr)

    log_pipeline_run(
        session,
        "success",
        loaded_financials,
        metadata={
            "file": filepath.name,
            "pnl_rows": len(pnl_rows),
            "bs_rows": len(bs_rows),
            "cf_rows": len(cf_rows),
            "periods": len(periods),
        },
    )

    print(f"\nDone. {loaded_financials} rows loaded from {filepath.name}.", file=sys.stderr)


if __name__ == "__main__":
    main()
